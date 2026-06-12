"""
Router Export Excel — activités et journal de ferme.
Préfixe : /api/app
  GET /activites/export-excel          → Excel toutes activités org
  GET /parcelles/{id}/activites/export-excel → Excel activités d'une parcelle
"""
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import current_user
from app.models import User
from app.models.ferme import Activite, StatutActivite
from app.models.champ import Parcelle

router = APIRouter(prefix="/api/app", tags=["Export Excel"])

_STATUT_LABELS = {
    StatutActivite.PLANIFIE:  "Planifié",
    StatutActivite.EN_COURS:  "En cours",
    StatutActivite.TERMINE:   "Réalisé",
    StatutActivite.ANNULE:    "Annulé",
}


def _build_excel(activites: list, titre: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Activités"

    vert_fonce = "166534"
    vert_clair = "DCFCE7"
    gris_clair = "F9FAFB"

    # Titre
    ws.merge_cells("A1:H1")
    ws["A1"] = titre
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=vert_fonce)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sous-titre
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y')} — AgroScan Pro"
    ws["A2"].font = Font(italic=True, size=10, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # En-têtes colonnes
    headers = ["ID", "Parcelle", "Type d'activité", "Description", "Date prévue", "Date réelle", "Statut", "Créé le"]
    widths  = [8,    20,         22,                 35,            14,            14,            12,       14]

    border_thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 20

    # Données
    statut_colors = {
        StatutActivite.PLANIFIE: "DBEAFE",
        StatutActivite.EN_COURS: "FEF9C3",
        StatutActivite.TERMINE:  vert_clair,
        StatutActivite.ANNULE:   "FEE2E2",
    }

    for row_idx, act in enumerate(activites, 4):
        bg = statut_colors.get(act.statut, "FFFFFF")
        fill = PatternFill("solid", fgColor=bg) if row_idx % 2 == 0 else PatternFill("solid", fgColor=gris_clair)

        values = [
            act.id,
            act._parcelle_nom or "—",
            act.type_activite or "—",
            act.description or "—",
            str(act.date_prevue) if act.date_prevue else "—",
            "—",
            _STATUT_LABELS.get(act.statut, str(act.statut)),
            str(act.created_at.date()) if act.created_at else "—",
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 4))
        ws.row_dimensions[row_idx].height = 18

    # Figer la ligne d'en-tête
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _enrich_parcelle_noms(activites: list, db: Session, org_id: int):
    ids = {a.parcelle_id for a in activites if a.parcelle_id}
    noms = {}
    if ids:
        rows = db.query(Parcelle.id, Parcelle.nom).filter(Parcelle.id.in_(ids)).all()
        noms = {r.id: r.nom for r in rows}
    for a in activites:
        a._parcelle_nom = noms.get(a.parcelle_id, "—")
    return activites


@router.get("/activites/export-excel")
def exporter_activites_excel(
    statut: str = Query(None, description="Filtrer par statut : planifie|en_cours|termine|annule"),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Exporte toutes les activités de l'organisation en Excel."""
    q = db.query(Activite).filter(Activite.org_id == user.org_id)
    if statut:
        try:
            q = q.filter(Activite.statut == StatutActivite(statut.upper()))
        except ValueError:
            pass
    activites = q.order_by(Activite.date_prevue.asc()).all()
    _enrich_parcelle_noms(activites, db, user.org_id)

    titre = f"Activités — AgroScan Pro"
    excel = _build_excel(activites, titre)
    filename = f"activites_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/parcelles/{parcelle_id}/activites/export-excel")
def exporter_activites_parcelle_excel(
    parcelle_id: int,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Exporte les activités d'une parcelle en Excel."""
    from fastapi import HTTPException
    p = db.query(Parcelle).filter_by(id=parcelle_id, org_id=user.org_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Parcelle introuvable.")

    activites = (
        db.query(Activite)
        .filter_by(parcelle_id=parcelle_id, org_id=user.org_id)
        .order_by(Activite.date_prevue.asc())
        .all()
    )
    for a in activites:
        a._parcelle_nom = p.nom

    titre = f"Activités — {p.nom}"
    excel = _build_excel(activites, titre)
    filename = f"activites_{p.code_parcelle or parcelle_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
